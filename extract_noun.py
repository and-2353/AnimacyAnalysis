from openpyxl import Workbook
from openpyxl import load_workbook
import csv
import nltk
from pprint import pprint

def extract_nouns_from_NGSL_by_rule():
    """
    NGSL上にある単語のうち、語形変化が1つだけ載ってるものを名詞とみなして抽出, csv出力。
    """
    wb = load_workbook('data/NGSL+1.01+by+band.xlsx')
    with open('data/nouns_not_filtered.csv', 'w', newline="") as f:
        writer = csv.writer(f)
        for sheet_name in ['1st 1000', '2nd 1000', '3rd 1000']:
            ws = wb[sheet_name]
            for i in range(1, 1001):
                lemma = ws[f'a{i}'].value
                infl_0 = ws[f'c{i}'].value
                infl_1 = ws[f'd{i}'].value
                if infl_0 != None and infl_1 == None:
                    writer.writerow([lemma])

def extract_nouns_from_NGSL_by_NLTK():
    """
    NGSL上にある単語のうち、単語をNLTKによる品詞推定の入力とし、以下の形式でcsv出力。
    単語, 品詞タグ, 品詞名(日本語) 
    """
    pos_ja = {
        'CC':'調整接続詞',
        'CD':'基数',
        'DT':'限定詞',
        'EX':'存在を表すthere',
        'FW':'外国語',
        'IN':'前置詞または従属接続詞',
        'JJ':'形容詞',
        'JJR':'形容詞 (比較級)',
        'JJS':'形容詞 (最上級)',
        'LS':'-',
        'MD':'法',
        'NN':'名詞',
        'NNS':'名詞 (複数形)',
        'NNP':'固有名詞',
        'NNPS':'固有名詞 (複数形)',
        'PDT':'前限定辞',
        'POS':'所有格の終わり',
        'PRP':'人称代名詞 (PP)',
        'PRP$':'所有代名詞 (PP$)',
        'RB':'副詞',
        'RBR':'副詞 (比較級)',
        'RBS':'副詞 (最上級)',
        'RP':'不変化詞',
        'SYM':'記号',
        'TO':'前置詞to',
        'UH':'感嘆詞',
        'VB':'動詞 (原形)',
        'VBD':'動詞 (過去形)',
        'VBG':'動詞 (動名詞または現在分詞)',
        'VBN':'動詞 (過去分詞)',
        'VBP':'動詞 (三人称単数以外の現在形)',
        'VBZ':'動詞 (三人称単数の現在形)',
        'WDT':'Wh 限定詞',
        'WP':'wh 代名詞',
        'WP$':'所有 Wh 代名詞',
        'WRB':'Wh 副詞'
    }
    nltk.download('averaged_perceptron_tagger')
    wb = load_workbook('data/NGSL+1.01+by+band.xlsx')
    
    with open('data/pos_of_words_in_NGSL.csv', 'w', newline="") as f:
        writer = csv.writer(f)
        for sheet_name in ['1st 1000', '2nd 1000', '3rd 1000']:
            ws = wb[sheet_name]
            for i in range(1, 1001):
                lemma = ws[f'a{i}'].value
                if lemma == None:
                    continue
                pos = nltk.pos_tag([lemma])[0][1]
                ja = pos_ja[pos] 
                writer.writerow([lemma, pos, ja])


def filter_nouns_hard():
    """
    nouns_not_filtered.csv の各語の品詞をpos_of_words_in_NGSL.csv から取得し、
    「名詞 in 品詞名」であるものをcsv出力。
    """
    lemma_and_pos = {}
    with open('data/pos_of_words_in_NGSL.csv') as f:
        reader = csv.reader(f)
        for row in reader:
            lemma_and_pos[row[0]] =  row[2]
    filtered_nouns = []
    with open('data/nouns_not_filtered.csv') as f:
        reader_ = csv.reader(f)
        for row in reader_:
            lemma = row[0]
            pos = lemma_and_pos[lemma]
            if '名詞' in pos:
                filtered_nouns.append(lemma)
    with open('nouns/nouns_v3.csv', 'w', newline="") as f:
        writer = csv.writer(f)
        for item in filtered_nouns:
            writer.writerow([item])

def filter_nouns_soft():
    """
    nouns_not_filtered.csv の各語の品詞をpos_of_words_in_NGSL.csv から取得し、
    「形容詞 not in 品詞名」であるものをcsv出力。
    """
    lemma_and_pos = {}
    with open('data/pos_of_words_in_NGSL.csv') as f:
        reader = csv.reader(f)
        for row in reader:
            lemma_and_pos[row[0]] =  row[2]
    filtered_nouns = []
    with open('data/nouns_not_filtered.csv') as f:
        reader_ = csv.reader(f)
        for row in reader_:
            lemma = row[0]
            pos = lemma_and_pos[lemma]
            if '形容詞' not in pos:
                filtered_nouns.append(lemma)
    with open('nouns/nouns_v4.csv', 'w', newline="") as f:
        writer = csv.writer(f)
        for item in filtered_nouns:
            writer.writerow([item])

def compare_soft_hard_filter():
    """
    filter_nouns_soft() と filter_nouns_hard() の出力結果を比較
    差分要素とその数を表示
    """
    l_soft = []
    with open('nouns/nouns_v4.csv') as f:
        reader = csv.reader(f)
        for row in reader:
            l_soft.append(row[0])
    l_hard = []
    with open('nouns/nouns_v3.csv') as f:
        reader = csv.reader(f)
        for row in reader:
            l_hard.append(row[0])
    diff_list = set(l_soft) ^ set(l_hard)
    print(diff_list)
    print(len(diff_list))

def divert_label():
    """
    別のファイルにつけたラベルを転用する
    """
    nouns_list_from = {}
    with open('nouns/nouns_v8.csv') as f: # ラベル付け済みファイル ラベル参照先
        reader = csv.reader(f)
        _ = next(reader) # skip headline
        for row in reader:
            lemma, label = row
            nouns_list_from[lemma] = label
    nouns_list_to = []
    with open('nouns/nouns_v9.csv') as f: # 未ラベル付けファイル 単語参照先
        reader = csv.reader(f)
        for row in reader:
            lemma = row[0]
            if lemma in nouns_list_from:
                nouns_list_to.append([lemma, nouns_list_from[lemma]])
            else:
                nouns_list_to.append([lemma, ''])
    with open('nouns/nouns_v10.csv', 'w', newline="") as f: # 出力先
        writer = csv.writer(f)
        writer.writerow(['lemma', 'animacy'])
        for item in nouns_list_to:
            writer.writerow(item)

def arrange_label():
    """
    つけたラベルを整理する
        2を0にするファイル(nouns_v8.1.csv)
        2を1にするファイル(nouns_v8.2.csv)を作る
        3は除去する
    """
    nouns_list_from = {}
    with open('nouns/nouns_v8.csv') as f:
        reader = csv.reader(f)
        headline = next(reader) 
        with open('nouns/nouns_v8/nouns_v8.1.csv', 'w', newline="") as w1:
            with open('nouns/nouns_v8/nouns_v8.2.csv', 'w', newline="") as w2:
                writer1 = csv.writer(w1)
                writer2 = csv.writer(w2)
                writer1.writerow(headline)
                writer2.writerow(headline)      
                for row in reader:
                    lemma, label = row
                    if label in ("0", "1"):
                        writer1.writerow([lemma, label])
                        writer2.writerow([lemma, label])
                    elif label == "2":
                        writer1.writerow([lemma, 0])
                        writer2.writerow([lemma, 1])
                    elif label == "3":
                        continue
                    else:
                        print('Exception')

def count_label():
    """
    各ラベルの語数を数える
    """
    with open('nouns/nouns_v8/nouns_v8.2.csv') as f:
        reader = csv.reader(f)
        _ = next(reader) # skip headline
        count = {}
        for row in reader:
            lemma, label = row
            if label in count:
                count[label] += 1
            else:
                count[label] = 1
                print(label, lemma)
        print(count)

def extract_nouns_from_BNC():
    with open('data/lemma.num.txt') as f:
        lines = f.readlines()
        with open('nouns/nouns_v9.csv', 'w', newline="") as w:
            writer = csv.writer(w)
            for line in lines:
                line = line.split()
                lemma = line[2]
                if 'n' in line:
                    writer.writerow([lemma])

        

if __name__ == '__main__':
    #divert_label()
    #count_label()
    #arrange_label()
    #extract_nouns_from_BNC()
    divert_label()