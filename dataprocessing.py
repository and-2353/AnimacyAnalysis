import csv
import nltk
import gensim
import pickle
from pprint import pprint
import numpy as np
import random
import openpyxl

def arrange_label(file_r, file_w1, file_w2):
    """
    手動でつけたラベルを整理する
        2を0にするファイル(csv形式, file_w1)
        2を1にするファイル(csv形式, file_w2)を作る
        3は除去する
    """
    with open(file_r) as f:
        reader = csv.reader(f)
        headline = next(reader) 
        with open(file_w1, 'w', newline="") as w1:
            with open(file_w2, 'w', newline="") as w2:
                writer1 = csv.writer(w1)
                writer2 = csv.writer(w2)
                writer1.writerow(headline)
                writer2.writerow(headline)      
                for row in reader:
                    lemma, label = row
                    if label in ("0", "1"):
                        writer1.writerow([lemma, label])
                        writer2.writerow([lemma, label])
                    elif label == "2":
                        writer1.writerow([lemma, 0])
                        writer2.writerow([lemma, 1])
                    elif label == "3":
                        continue
                    else:
                        print('Exception')

def count_label(file_r):
    """
    各ラベルの語数を数える, 読み込むcsvファイルにheadlineがあるか確認する
    """
    with open(file_r) as f:
        reader = csv.reader(f)
        #_ = next(reader) # skip headline
        count = {}
        for row in reader:
            lemma = row[0] 
            label = row[-1]
            if label in count:
                count[label] += 1
            else:
                count[label] = 1
                print(label, lemma)
        print(count)

def arrange_data_to_discriminant_analysis(file_em, file_r, file_w):
    """
    ファイルにembeddingを結合する。
        csv(lemma,animacy 形式)を入力とし、
        csv(lemma,d0~d299,animacy)形式を出力する。
    読み込むcsvファイルにheadlineがあるか確認する
    """
    keyerror_num = 0
    with open(file_em, 'rb') as em:
        model = pickle.load(em)
        with open(file_w, 'w', newline="") as f_w:
            writer = csv.writer(f_w)
            with open(file_r) as f_r:
                reader = csv.reader(f_r)
                #_ = next(reader) # skip headline
                header = create_headline()
                writer.writerow(header)
                for row in reader:
                    try:
                        lemma = row[0]
                        animacy = row[1]
                        row_ = [lemma]
                        row_.extend(model[lemma])
                        row_.append(animacy)
                        writer.writerow(row_)
                    except KeyError:
                        print(f"{row[0]}: KeyError")
                        keyerror_num += 1
                        continue
    print('keyerror_num:', keyerror_num)


def create_headline():
    """
    arrange_data_to_discriminant_analysis 内で呼ばれ、
    出力するファイルのheadline(要素数302のlist)を作成する
    """
    header = ['lemma']
    for i in range(300):
        header.append('d'+str(i))
    header.append('animacy')
    return header

def random_labeling(file_r, file_w, n_zero, n_one):
    """
    手作業でつけたラベルをランダムなラベルに変更(ランダムラベリング)
    """
    zeros = [0 for i in range(n_zero)]
    ones = [1 for i in range(n_one)]
    labels = zeros + ones
    random.shuffle(labels)

    with open(file_w, 'w', newline="") as f: # 出力先
        writer = csv.writer(f)
        with open(file_r) as f_r:
            reader = csv.reader(f_r)
            _ = next(reader) # skip headline
            #assert len(labels) == len(reader)
            for label, row in zip(labels, reader):
                lemma = row[0]
                writer.writerow([lemma, label])

def remove_no_embedding_lemma(file_em, file_r, file_w):
    """
    embedding.pickle の見出し語として登録されていない語, 手作業ラベル=3の語を取り除く。
        csv(lemma,animacy) 形式を入力とし、
        csv(lemma,animacy) 形式を出力する。
    """
    keyerror_num = 0
    with open(file_em, 'rb') as em:
        model = pickle.load(em)
        with open(file_w, 'w', newline="") as f_w:
            writer = csv.writer(f_w)
            with open(file_r) as f_r:
                reader = csv.reader(f_r)        
                for row in reader:
                    try:
                        lemma = row[0]
                        animacy = row[1]
                        _ = model[lemma] # model[lemma] がなければ except へ
                        if animacy == "3":
                            continue
                        writer.writerow(row)
                    except KeyError:
                        print(f"{row[0]}: KeyError")
                        keyerror_num += 1
                        continue
    print('keyerror_num:', keyerror_num)

def extract_animate_and_middle_words(file_r, file_w):
    """
    有生性ラベルが 有生性あり or 迷う(1 or 2) の語を抽出する。
        入力:csv(lemma, animacy の2列, headlineなしを想定)
        出力:xlsx(seat:[1, 2](有生性ラベルを示す), 各シートlemma, animacyの2列)
    """
    extracted_nums = {1:2, 2:2} # SheetName: RowNo.
    wb = openpyxl.Workbook()
    wb.worksheets[0].title = "1"
    wb.create_sheet(title="2")
    with open(file_r) as f_r:
        reader = csv.reader(f_r)        
        for row in reader:
            lemma = row[0]
            animacy = int(row[1])
            if animacy in (1, 2):
                ws = wb[str(animacy)]
                row = str(extracted_nums[animacy])
                ws["A"+row] = lemma
                ws["B"+row] = animacy
                extracted_nums[animacy] += 1
    wb.save(file_w)

def extract_words_by_label(file_r, file_w, label):
    """
    有生性ラベルが label の語を抽出する。
        入力:csv(lemma, animacy の2列, headlineなしを想定)
        出力:csv(lemma, animacy の2列, headlineなしを想定)
    """
    extracted_num = 0
    with open(file_w, 'w', newline="") as f_w:
        writer = csv.writer(f_w)
        with open(file_r) as f_r:
            reader = csv.reader(f_r)        
            for row in reader:
                lemma, animacy = row
                if animacy == str(label):
                    writer.writerow(row)
                    extracted_num += 1
    print('抽出語数:', extracted_num)
    

def arrange_balanced_data(file_ani, file_inani, file_w):
    ani_words = []
    with open(file_ani) as ani:
        reader = csv.reader(ani)        
        for row in reader:
            ani_words.append(row)
    n_ani_words = len(ani_words)

    inani_words = []
    with open(file_inani) as inani:
        reader_ = csv.reader(inani)
        for row in reader_:
            inani_words.append(row)
    inani_words_squeezed = random.sample(inani_words, n_ani_words)
    balanced_data = ani_words + inani_words_squeezed
    random.shuffle(balanced_data)
    print("num of words in balanced_data:", len(balanced_data))

    with open(file_w, 'w', newline="") as f_w:
        writer = csv.writer(f_w)
        for row in balanced_data:
            writer.writerow(row)

def extract_words_by_labels(file_r, file_w, labels):
    """
    有生性ラベルが labels の語を抽出する。
        入力:csv(lemma, animacy の2列, headlineなしを想定)
        出力:csv(lemma, animacy の2列, headlineなしを想定)
    """
    extracted_num = 0
    with open(file_w, 'w', newline="") as f_w:
        writer = csv.writer(f_w)
        with open(file_r) as f_r:
            reader = csv.reader(f_r)        
            for row in reader:
                lemma, animacy = row
                if animacy in labels:
                    writer.writerow(row)
                    extracted_num += 1
    print('抽出語数:', extracted_num)

def extract_words_by_attribute(xlsx_r, csv_w):
    """
    有生性ラベル=2の語のうち、特定の属性（ex.植物）の語を抽出。
    if文の条件属性によって変えないといけないので注意...
    """
    extracted_num = 0
    wb = openpyxl.load_workbook(xlsx_r)
    with open(csv_w, 'w', newline="") as f:
        writer = csv.writer(f)
        ws = wb['2']
        for i in range(2, 264):
            lemma = ws[f'a{i}'].value
            animacy = ws[f'b{i}'].value
            collective = ws[f'f{i}'].value
            plant = ws[f'g{i}'].value
            spirit = ws[f'h{i}'].value
            micro = ws[f'i{i}'].value
            special = ws[f'j{i}'].value
            if plant != None:
                writer.writerow([lemma, animacy])
            


if __name__ == '__main__':
    file_em = 'data/embedding/embedding.pickle'
    # file_r = 'extracted_nouns/BNC/nouns_bnc+lbl+key.csv'
    
    # file_w = 'extracted_nouns/BNC/nouns_bnc+lbl+key+bld(3).csv'
    # file_ww = 'extracted_nouns/BNC/nouns_bnc+lbl+key+bld(3)+em.csv'
    # file_r = 'extracted_nouns/BNC/nouns_bnc+lbl+key.csv'
    # files = ['extracted_nouns/BNC/nouns_bnc+lbl+key+rl(1=369)(2).csv', 'extracted_nouns/BNC/nouns_bnc+lbl+key+rl(1=369)(3).csv']
    # files_ = ['extracted_nouns/BNC/nouns_bnc+lbl+key+rl(1=369)(2)+em.csv', 'extracted_nouns/BNC/nouns_bnc+lbl+key+rl(1=369)(3)+em.csv']
    # n_zero, n_one = 2839, 369
    # for i in range(2):
    #     file_w = files[i]
    #     file_ww = files_[i]
    #     random_labeling(file_r, file_w, n_zero, n_one)
    #     arrange_data_to_discriminant_analysis(file_em, file_w, file_ww)
    # file_w = 'extracted_nouns/BNC/nouns_bnc+lbl+key+bld+em.csv'
    
    # extract_animate_words(file_r, file_w)
    # extract_words_by_label(file_r, file_w1, 1)
    # extract_words_by_label(file_r, file_w2, 0)
    #arrange_balanced_data(file_ani, file_inani, file_w)
    # file_r = 'extracted_nouns/BNC/nouns_bnc+lbl+key+ani_mid+spirit.csv'
    # file_w = 'extracted_nouns/BNC/nouns_bnc+lbl+key+ani_mid+spirit+em.csv'
    # file_r = 'extracted_nouns/BNC/nouns_bnc+lbl+key+bi3.csv'
    # file_w = 'extracted_nouns/BNC/nouns_bnc+lbl+key+bi3+em.csv'
    
    # xlsx_r = 'extracted_nouns/BNC/nouns_bnc+lbl+key+ani_mid.xlsx'
    # csv_w = 'extracted_nouns/BNC/nouns_bnc+lbl+key+ani_mid+plant.csv'
    # file_ww = 'extracted_nouns/BNC/nouns_bnc+lbl+key+ani_mid+plant+em.csv'
    # extract_words_by_attribute(xlsx_r, csv_w)
    # arrange_data_to_discriminant_analysis(file_em, csv_w, file_ww)
    
    # file_r = 'extracted_nouns/BNC/nouns_bnc+lbl+key.csv'
    
    # labels = ['0', '1']
    # extract_words_by_labels(file_r, file_w, labels)
    # file_ani = 'extracted_nouns/BNC/nouns_bnc+lbl+key+ani.csv'
    # file_inani = 'extracted_nouns/BNC/nouns_bnc+lbl+key-ani.csv'
    for i in range(4, 11):
        file_r = f'extracted_nouns/BNC/nouns_bnc+lbl+key+bld({i}).csv'
        file_w = f'extracted_nouns/BNC/nouns_bnc+lbl+key+bld({i})+em.csv'
        arrange_data_to_discriminant_analysis(file_em, file_r, file_w)
        # arrange_balanced_data(file_ani, file_inani, file_w)